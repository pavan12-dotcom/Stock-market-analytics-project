import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import os

BASE = '/home/claude/stock_project'
df      = pd.read_csv(f'{BASE}/data/stock_history.csv', parse_dates=['date'])
stats   = pd.read_csv(f'{BASE}/data/stock_stats.csv')
snap    = pd.read_csv(f'{BASE}/data/latest_snapshot.csv')

wb = Workbook()

# ── helpers ────────────────────────────────────────────────────────
def hf(color): return PatternFill('solid', start_color=color, end_color=color)
def bf(sz=10,bold=False,color='1F2937'): return Font(name='Arial',size=sz,bold=bold,color=color)
def hdr_font(sz=10): return Font(name='Arial',size=sz,bold=True,color='FFFFFF')
def thin(): 
    s = Side(style='thin',color='CCCCCC')
    return Border(left=s,right=s,top=s,bottom=s)
def set_w(ws,col,w): ws.column_dimensions[get_column_letter(col)].width = w
def center(c): c.alignment = Alignment(horizontal='center',vertical='center')

def title_row(ws,text,cols=16,bg='0D1117',fg='58A6FF'):
    for c in range(1,cols+1): ws.cell(1,c).fill = hf(bg)
    ws.merge_cells(f'A1:{get_column_letter(cols)}1')
    t = ws['A1']; t.value = text
    t.font = Font(name='Arial',size=14,bold=True,color=fg)
    t.alignment = Alignment(horizontal='center',vertical='center')
    ws.row_dimensions[1].height = 30

# ══════════════════════════════════════════════════════════════════
# SHEET 1: Market Overview
# ══════════════════════════════════════════════════════════════════
ws = wb.active; ws.title = 'Market Overview'
ws.sheet_view.showGridLines = False
title_row(ws,'📈  REAL-TIME STOCK MARKET ANALYTICS PLATFORM  |  10 Stocks · 1-Year Analysis')

# KPI strip
kpi_bg = ['1D4ED8','15803D','7C3AED','B45309','991B1B','0E7490']
kpi_labels = ['MSFT (Best Sharpe)','WMT (Lowest Vol)','NVDA (Top Momentum)','Avg Daily Volume','Tickers Tracked','Data Points']
avg_vol = df.groupby('ticker')['volume'].mean().mean() if 'volume' in df.columns else 0
kpi_vals = ['$509  ▲1.81x','$74.85  σ11.7%','$817  +3.2%',f'{avg_vol/1e6:.1f}M','10','2,610']

for i,(lbl,val,bg) in enumerate(zip(kpi_labels,kpi_vals,kpi_bg)):
    col = i*3+1
    ws.merge_cells(start_row=3,start_column=col,end_row=3,end_column=col+2)
    ws.merge_cells(start_row=4,start_column=col,end_row=4,end_column=col+2)
    ws.merge_cells(start_row=5,start_column=col,end_row=5,end_column=col+2)
    l = ws.cell(3,col,lbl); l.font=Font(name='Arial',size=8,color='BFDBFE'); l.fill=hf(bg); center(l)
    v = ws.cell(4,col,val); v.font=Font(name='Arial',size=13,bold=True,color='FFFFFF'); v.fill=hf(bg); center(v)
    ws.cell(5,col,'').fill=hf(bg)
ws.row_dimensions[3].height=18; ws.row_dimensions[4].height=28; ws.row_dimensions[5].height=8

# Stock snapshot table
ws.cell(7,1,'LIVE STOCK SNAPSHOT').font=Font(name='Arial',size=11,bold=True,color='0D1117')
hdrs = ['Ticker','Company','Sector','Last Price','Change','Change %','52W High','52W Low','Ann. Return %','Ann. Vol %','Sharpe']
for i,h in enumerate(hdrs,1):
    c = ws.cell(8,i,h); c.font=hdr_font(9); c.fill=hf('0D1117'); center(c); c.border=thin()

merged = stats.merge(snap[['ticker','name','sector','w52h','w52l']],on='ticker',how='left') if 'name' in snap.columns else stats
for ri,row in merged.iterrows():
    r = ri+9
    chg = row.get('change_pct',0)
    chg_color = '16A34A' if chg >= 0 else 'DC2626'
    row_bg = 'F8FAFC' if ri%2==0 else 'FFFFFF'
    vals = [
        row['ticker'],
        row.get('name', row['ticker']),
        row.get('sector','N/A'),
        f"${row['last_price']:.2f}",
        f"{'+' if chg>=0 else ''}{chg*row['last_price']/100:.2f}",
        f"{'+' if chg>=0 else ''}{chg:.2f}%",
        f"${row.get('w52h',0):.2f}" if row.get('w52h',0) else 'N/A',
        f"${row.get('w52l',0):.2f}" if row.get('w52l',0) else 'N/A',
        f"{row.get('ann_return',0):.1f}%",
        f"{row.get('ann_vol',0):.1f}%",
        f"{row.get('sharpe',0):.2f}",
    ]
    for ci,v in enumerate(vals,1):
        c = ws.cell(r,ci,v); c.border=thin(); center(c); c.fill=hf(row_bg)
        if ci==5 or ci==6: c.font=Font(name='Arial',size=10,color=chg_color,bold=True)
        elif ci==10: c.font=Font(name='Arial',size=10,color='1F2937',bold=False)
        else: c.font=bf(10)

widths=[8,22,14,11,10,10,10,10,13,11,9]
for i,w in enumerate(widths,1): set_w(ws,i,w)

# ══════════════════════════════════════════════════════════════════
# SHEET 2: Risk & Performance
# ══════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('Risk & Performance')
ws2.sheet_view.showGridLines = False
title_row(ws2,'⚡  RISK & PERFORMANCE ANALYTICS')

hdrs2 = ['Ticker','Ann. Return %','Ann. Volatility %','Sharpe Ratio','Beta','Risk Level','Performance']
for i,h in enumerate(hdrs2,1):
    c=ws2.cell(3,i,h); c.font=hdr_font(10); c.fill=hf('0D1117'); center(c); c.border=thin()

risk_colors = {'Low':'15803D','Medium':'B45309','High':'DC2626','Very High':'7F1D1D'}
perf_colors = {'Strong':'15803D','Moderate':'1D4ED8','Weak':'B45309','Negative':'DC2626'}

for ri,row in stats.iterrows():
    r=ri+4
    beta  = row.get('beta',1.0) if not pd.isna(row.get('beta',np.nan)) else 1.0
    vol   = row.get('ann_vol',20)
    ret   = row.get('ann_return',0)
    shrp  = row.get('sharpe',0)
    risk  = 'Very High' if vol>40 else 'High' if vol>25 else 'Medium' if vol>15 else 'Low'
    perf  = 'Strong' if ret>20 else 'Moderate' if ret>5 else 'Weak' if ret>0 else 'Negative'
    row_bg = 'F8FAFC' if ri%2==0 else 'FFFFFF'
    vals=[row['ticker'],f"{ret:.1f}%",f"{vol:.1f}%",f"{shrp:.2f}",f"{beta:.2f}",risk,perf]
    for ci,v in enumerate(vals,1):
        c=ws2.cell(r,ci,v); c.border=thin(); center(c); c.fill=hf(row_bg); c.font=bf(10)
        if ci==6: c.font=Font(name='Arial',size=10,bold=True,color=risk_colors.get(risk,'333333'))
        if ci==7: c.font=Font(name='Arial',size=10,bold=True,color=perf_colors.get(perf,'333333'))

for i,w in enumerate([10,16,18,13,10,12,13],1): set_w(ws2,i,w)

# ══════════════════════════════════════════════════════════════════
# SHEET 3: Monthly Returns
# ══════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('Monthly Returns')
ws3.sheet_view.showGridLines = False
title_row(ws3,'📅  MONTHLY RETURN ANALYSIS (%)')

df['year_month'] = df['date'].dt.to_period('M').astype(str)
monthly_ret = df.groupby(['ticker','year_month']).apply(
    lambda x: round((x['close'].iloc[-1]/x['close'].iloc[0]-1)*100,2)
).reset_index(name='ret')
heat = monthly_ret.pivot(index='ticker',columns='year_month',values='ret').fillna(0)
months = sorted(heat.columns)

ws3.cell(3,1,'Ticker').font=hdr_font(10); ws3.cell(3,1).fill=hf('0D1117'); ws3.cell(3,1).border=thin(); center(ws3.cell(3,1))
for j,m in enumerate(months,2):
    c=ws3.cell(3,j,m); c.font=hdr_font(8); c.fill=hf('0D1117'); c.border=thin(); center(c)
avg_col = len(months)+2
ws3.cell(3,avg_col,'Avg %').font=hdr_font(9); ws3.cell(3,avg_col).fill=hf('0D1117'); ws3.cell(3,avg_col).border=thin(); center(ws3.cell(3,avg_col))

for ri,ticker in enumerate(heat.index):
    r=ri+4
    c=ws3.cell(r,1,ticker); c.font=Font(name='Arial',size=10,bold=True,color='0D1117'); c.border=thin(); center(c)
    row_vals=[]
    for j,m in enumerate(months,2):
        val = heat.loc[ticker,m] if m in heat.columns else 0
        row_vals.append(val)
        cell=ws3.cell(r,j,round(val,2))
        # Color gradient: green=positive, red=negative
        intensity = min(int(abs(val)*8),200)
        if val>0: bg=f'00{255-intensity:02X}00' if intensity<200 else '006400'
        elif val<0: bg=f'{min(255,155+intensity):02X}0000'
        else: bg='F8FAFC'
        cell.fill=hf(bg)
        cell.font=Font(name='Arial',size=9,color='FFFFFF' if abs(val)>5 else '1F2937')
        cell.border=thin(); center(cell)
    avg_val = round(np.mean(row_vals),2)
    ac=ws3.cell(r,avg_col,avg_val)
    ac.font=Font(name='Arial',size=10,bold=True,color='15803D' if avg_val>=0 else 'DC2626')
    ac.border=thin(); center(ac); ac.fill=hf('F0FDF4' if avg_val>=0 else 'FEF2F2')

ws3.column_dimensions['A'].width=9
for j in range(2,len(months)+3): ws3.column_dimensions[get_column_letter(j)].width=9

# ══════════════════════════════════════════════════════════════════
# SHEET 4: Charts
# ══════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet('Analysis Charts')
ws4.sheet_view.showGridLines = False
title_row(ws4,'📊  VISUAL ANALYSIS CHARTS',cols=20)

chart_files = [
    ('chart1_normalised_performance.png', 3,  1, 'Normalised Price Performance'),
    ('chart2_monthly_heatmap.png',        32, 1, 'Monthly Returns Heatmap'),
    ('chart3_risk_return.png',            60, 1, 'Risk-Return Scatter'),
    ('chart4_correlation.png',            60, 11,'Correlation Matrix'),
    ('chart5_price_volume.png',           90, 1, 'Price & Volume — Top 4'),
    ('chart6_rolling_sharpe.png',         120,1, 'Rolling Sharpe Ratio'),
]
for fname,row,col,label in chart_files:
    fpath = f'{BASE}/charts/{fname}'
    if os.path.exists(fpath):
        ws4.cell(row,col,label).font=Font(name='Arial',size=11,bold=True,color='58A6FF')
        img=XLImage(fpath); img.width=680; img.height=310
        ws4.add_image(img,f'{get_column_letter(col)}{row+1}')
ws4.column_dimensions['A'].width=90

# ══════════════════════════════════════════════════════════════════
# SHEET 5: Raw Data
# ══════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet('Raw Data')
ws5.sheet_view.showGridLines = False
title_row(ws5,'📋  RAW STOCK DATA (Historical OHLCV)')

raw_cols = [c for c in ['ticker','name','sector','date','open','high','low','close','volume','daily_ret'] if c in df.columns]
for i,h in enumerate(raw_cols,1):
    c=ws5.cell(3,i,h.replace('_',' ').title()); c.font=hdr_font(9); c.fill=hf('0D1117'); c.border=thin(); center(c)

for ri,row in df[raw_cols].iterrows():
    r=ri+4
    for ci,col in enumerate(raw_cols,1):
        val=row[col]
        if col=='date': val=str(val)[:10]
        elif col in ['open','high','low','close']: val=round(float(val),2)
        elif col=='daily_ret' and pd.notna(val): val=round(float(val)*100,3)
        c=ws5.cell(r,ci,val); c.font=bf(8); c.border=thin(); center(c)
        c.fill=hf('F8FAFC') if ri%2==0 else hf('FFFFFF')

for i,w in enumerate([8,20,12,12,9,9,9,9,14,12],1): set_w(ws5,i,w)

# Save
path = f'{BASE}/reports/Stock_Market_Analytics.xlsx'
os.makedirs(f'{BASE}/reports', exist_ok=True)
wb.save(path)
print(f"✓ Excel report saved: {path}")
